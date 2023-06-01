from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
import openpyxl
import datetime
import time

# get current day
day_index = datetime.datetime.today().weekday()
days = ["Saturday", "Sunday", "Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]
current_day = datetime.datetime.now().strftime('%A')
print("Today is '{}'".format(current_day))

# load worksheet
workbook = openpyxl.load_workbook('file.xlsx')
worksheet = workbook[current_day]

# find keywords list
keywords = []
for row in worksheet.iter_rows(min_row=3, min_col=3, max_col=3):
    for cell in row:
        keywords.append(cell.value)
print("Keywords are Collected Successfully: ", keywords, "\n")

# search for each keyword in Google and find longest and shortest option
driver = webdriver.Chrome()
row_num = 3
for keyword in keywords:
    driver.get("https://www.google.com/")
    search_box = driver.find_element(By.NAME, "q")
    search_box.send_keys(keyword)
    time.sleep(1)  # Wait to load suggestions
    search_box.send_keys(Keys.ARROW_DOWN)  # Move the cursor down to trigger every suggestion

    # Searching the elements
    suggestion_elements = driver.find_elements(By.CSS_SELECTOR, "li.sbct")

    # Searching the suggestions
    suggestion_list = [suggestion.text.split('\n')[0] for suggestion in suggestion_elements if
                       len(suggestion.text) > 0]  # Split to take the text only
    print("Keyword: {}\nSuggestions are: {}".format(keyword, suggestion_list))

    longest_suggestion = max(suggestion_list, key=len)  # Finding the longest one
    shortest_suggestion = min(suggestion_list, key=len)  # Fonding the shortest one
    print("Longest Suggestion is '{}'\nShortest Suggestion is '{}'\n".format(longest_suggestion, shortest_suggestion))

    # Store the values in the worksheet
    worksheet.cell(row=row_num, column=4).value = longest_suggestion
    worksheet.cell(row=row_num, column=5).value = shortest_suggestion

    row_num += 1

driver.quit()

# save changes to worksheet
workbook.save('file.xlsx')
print('Suggestions are Successfully Stored in the WorkSheet')
